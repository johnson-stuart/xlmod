import pandas as pd
import numpy as np
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import Alignment

#May need $ pip install xlwt openpyxl xlsxwriter xlrd

today = date.today()

wb = load_workbook('testsheet.xlsx')
ws = wb.active
ws.insert_cols(3)

ws['C1'] = "Version No."
ws['C2'] = "v1 Sys Test {}".format(today.strftime("%d/%m/%Y"))
ws['C2'].alignment = Alignment(wrap_text = True)

wb.save('newsheet.xlsx')
